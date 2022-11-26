import socket,time,threading,queue,os,sys,datetime,schedule,psutil
from openpyxl import Workbook,load_workbook
from smtplib import SMTP
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

name_list=os.path.basename(__file__).split('.')
name_list[-1]="exe"
exe_name_list=[".".join(name_list),"_visible.".join(name_list)]
print(exe_name_list)
process_count=0
for p in psutil.process_iter():
    # print(p.name())
    if p.name() in exe_name_list:
        process_count=process_count+1
        # print
    if process_count>2:
        print(process_count)
        print("Already execution file ran")
        sys.exit() 

setting_file=open('msil_server_settings.txt','r')
print("MT: I am a Server of MSIL vision")

filedic={}
for line in setting_file:
    file_data=line.strip().split('===')
    a=file_data[0]
    b=file_data[1]
    filedic[a]=b
setting_file.close()
ipaddress_of_system=filedic.pop('ipaddress_of_system')
port_to_listen=int(filedic.pop('port_to_listen'))
shiftA_start=filedic.pop('shiftA_start_time')
shiftB_start=filedic.pop('shiftB_start_time')
shiftC_start=filedic.pop('shiftC_start_time')
filename_of_excel=filedic.pop('filename_of_excel_sheet')
mail_user_name=filedic.pop('mail_user_name')
app_password=filedic.pop('app_password')
prior_time=int(filedic.pop('prior_time_for_trigger_events_in_seconds'))
header_row=int(filedic.pop('header_row_in_excel'))
headers_list=filedic.pop('headers').strip().split(",")
date_column=int(filedic.pop('date_column_in_excel'))
shift_column=int(filedic.pop('shift_column_in_excel'))
time_column=int(filedic.pop('time_column_in_excel'))
elr_bc_data_column=int(filedic.pop('elr_bc_data_column_in_excel'))
sc_bc_data_column=int(filedic.pop('sc_bc_data_column_in_excel'))
vision_status_column=int(filedic.pop('vision_status_column_in_excel'))
scheduler_delay=int(filedic.pop('scheduler_delay_in_seconds'))
maximum_retry_count_excel=int(filedic.pop('maximum_retry_count_to_data_save_excel'))
maximum_retry_count_mail=int(filedic.pop('maximum_retry_count_trigger_mail'))
retry_delay=int(filedic.pop('retry_delay_in_milliseconds'))/1000

To_mail_pos_list=list(map(int,filedic.pop('To_mails_pos_in_mail_list').strip().split(",")))
Cc_mail_pos_list=list(map(int,filedic.pop('Cc_mails_pos_in_mail_list').strip().split(",")))

A=list(map(int,shiftA_start.strip().split(":")))
B=list(map(int,shiftB_start.strip().split(":")))
C=list(map(int,shiftC_start.strip().split(":")))

mail_file = open("mail.txt", "r")
list_of_mails= []
To_list_of_mails=[]
Cc_list_of_mails=[]
for line in mail_file:
  list_of_mails.append(line.strip())
for i in To_mail_pos_list:
    To_list_of_mails.append(list_of_mails[i])
for i in Cc_mail_pos_list:
    Cc_list_of_mails.append(list_of_mails[i])
mail_file.close()

excel_queue=queue.Queue()

def get_shift(ct):
    startA=datetime.time(A[0],A[1],A[2])
    startB=datetime.time(B[0],B[1],B[2])
    startC=datetime.time(C[0],C[1],C[2])
    if startA<=ct<startB:
        return 'A'
    elif startB<=ct<startC:
        return 'B'
    else:
        return 'C'

def multi_threaded_client(connection):
    # connection.send(str.encode('Server is working:'))
    while True:
        data = connection.recv(2048)
        # response = 'Server message: ' + data.decode('utf-8')
        # if not data:
            # break
        # connection.sendall(str.encode(response))
        excel_queue.put([data,time.time()])
        print(data, time.time())
    connection.close()

def move_excel():
    while True:
        s_data,time_ins=excel_queue.get()
        t=datetime.datetime.fromtimestamp(time_ins)
        count=0
        while count<=maximum_retry_count_excel:
            try:
                data_list=[]
                # print( len(s_data))
                current_time=datetime.time(t.hour,t.minute,t.second)
                if current_time<datetime.time(A[0],A[1],A[2]):
                    dp=1
                else:
                    dp=0
                date=t-datetime.timedelta(days=dp)
                sheet_name=(date.strftime("%b%Y")).upper()
                date=date.strftime("%d-%m-%Y")
                shift=get_shift(current_time)
                time_format=t.strftime("%I:%M:%S_%p")
                for i in range(0, len(s_data),85):
                    data=bytearray(s_data[i:i+85])
                    # print(data)
                    vision_status=data[0]
                    # print( len(data))
                    # print(data[2])
                    elr_bc_data=bytearray(data[3:3+data[2]]).decode()
                    # print(data[45])
                    sens_cov_bc_data=bytearray(data[45:45+data[44]]).decode()
                    data_list.append([date,shift,elr_bc_data,sens_cov_bc_data,vision_status,time_format])
                # print("test")
                # print(data_list)
                if not os.path.isfile(filename_of_excel):
                    wb=Workbook()
                else:
                    wb=load_workbook(filename_of_excel)
                    
                if not sheet_name in wb.sheetnames:
                    wb.create_sheet(sheet_name)
                    ws=wb[sheet_name]
                    for index_j,j in enumerate(headers_list):
                        ws.cell(row=header_row,column=index_j+1).value=j
                ws=wb[sheet_name]
                for i in data_list:
                    mr=ws.max_row
                    for index_j,j in enumerate(i):
                        ws.cell(row=mr+1,column=index_j+1).value=j            
                while True:
                    try:
                        wb.save(filename_of_excel)
                        print(f"EXH: Hourly and Idle time data saved successfully {time.strftime('%d-%m-%Y_%I.%M.%S_%p')}")
                    except Exception as e:
                        print(f"EXH: Data not saved, Close the excel file({filename_of_excel}) if it is opened. Retrying to save...\n Error: {e}")
                        time.sleep(3)
                        continue
                    break
                count=0
                wb.close()
                break
            except Exception as e:
                count=count+1
                print(e)
                time.sleep(retry_delay)
        
def make_content_and_send_summary():
    time_eve=round(time.time())-prior_time
    count=0
    while count<=maximum_retry_count_mail:
        try:
            t=datetime.datetime.fromtimestamp(time_eve)
            current_time=datetime.time(t.hour,t.minute,t.second)
            if current_time<datetime.time(A[0],A[1],A[2]):
                dp=1
            else:
                dp=0
            date=t-datetime.timedelta(days=dp)
            sheet_name=(date.strftime("%b%Y")).upper()
            date=date.strftime("%d-%m-%Y")
            shift=get_shift(current_time)
            # print(filename_of_excel)
            if os.path.isfile(filename_of_excel):
                wb=load_workbook(filename_of_excel)
                ws=wb[sheet_name]
                # xl_headers=[]
                required_row_list=[]
                mr=ws.max_row
                # print(mr)
                # for i in ws[header_row]:
                #         xl_headers.append(i.value)
                for i in range(1,mr+1):
                    date_ex=ws.cell(row=i,column=date_column).value
                    shift_ex=ws.cell(row=i,column=shift_column).value
                    if date_ex==date and shift_ex==shift :
                        required_row_list.append(i)
                # print(required_row_list)
                if not required_row_list==[]:
                    elr_data_list=[]
                    elr_count_list=[]
                    elr_vis_res_list=[]
                    elr_time_list=[]
                    sc_data_list=[]
                    sc_count_list=[]
                    sc_vis_res_list=[]
                    sc_time_list=[]
                    for i in required_row_list:
                        elr_data=ws.cell(row=i,column=elr_bc_data_column).value
                        sc_data=ws.cell(row=i,column=sc_bc_data_column).value
                        vis_res=int(ws.cell(row=i,column=vision_status_column).value)
                        time_data=ws.cell(row=i,column=time_column).value

                        if elr_data not in elr_data_list:
                            elr_data_list.append(elr_data)
                            elr_count_list.append(1)
                            elr_vis_res_list.append(vis_res)
                            elr_time_list.append(time_data)
                        else:
                            elr_index=elr_data_list.index(elr_data)
                            elr_count_list[elr_index]=elr_count_list[elr_index]+1
                            elr_time_list[elr_index]=time_data
                            if int(elr_vis_res_list[elr_index])!=1:
                                elr_vis_res_list[elr_index]=vis_res
                        if sc_data not in sc_data_list:
                            sc_data_list.append(sc_data)
                            sc_count_list.append(1)
                            sc_vis_res_list.append(vis_res)
                            sc_time_list.append(time_data)
                        else:
                            sc_index=sc_data_list.index(sc_data)
                            sc_count_list[sc_index]=sc_count_list[sc_index]+1
                            sc_time_list[sc_index]=time_data
                            if int(sc_vis_res_list[sc_index])!=1:
                                sc_vis_res_list[sc_index]=vis_res
                    max_elr_count=max(elr_count_list)
                    elr_ftr_c=elr_count_list.count(1)
                    sc_ftr_c=sc_count_list.count(1)
                    elr_ftr_p=round((elr_ftr_c/len(elr_count_list))*100,1)
                    sc_ftr_p=round((sc_ftr_c/len(sc_count_list))*100,1)
                    ftr_c=min(elr_ftr_c,sc_ftr_c)
                    ftr_p=min(elr_ftr_p,sc_ftr_p)
                    elr_ok_dic={}
                    sc_ok_dic={}
                    elr_nok_list=[]
                    sc_nok_list=[]
                    for co in range(max_elr_count,1,-1):
                        if co in elr_count_list:
                            indices = [i for i, x in enumerate(elr_count_list) if x == co]
                            for index in indices:
                                if elr_vis_res_list[index]!=0:
                                    if co not in elr_ok_dic:
                                        elr_ok_dic[co]=[]
                                    elr_ok_dic[co].append(elr_data_list[index])
                    max_sc_count=max(sc_count_list)
                    for co in range(max_sc_count,1,-1):
                        if co in sc_count_list:
                            indices = [i for i, x in enumerate(sc_count_list) if x == co]
                            for index in indices:
                                if sc_vis_res_list[index]!=0:
                                    if co not in sc_ok_dic:
                                        sc_ok_dic[co]=[]
                                    sc_ok_dic[co].append(sc_data_list[index])
                    # print(elr_ok_dic,sc_ok_dic)
                    elr_nok_indices = [i for i, x in enumerate(elr_vis_res_list) if x == 0]
                    sc_nok_indices = [i for i, x in enumerate(sc_vis_res_list) if x == 0]
                    for i in elr_nok_indices:
                        elr_nok_list.append(str(elr_data_list[i])+'<label style="background-color: aqua;"> ['+str(elr_count_list[i])+']</label>'+'<label style="background-color: #DFFF00;"> ['+str(elr_time_list[i])+']</label>')
                    for i in sc_nok_indices:
                        sc_nok_list.append(str(sc_data_list[i])+'<label style="background-color: aqua;"> ['+str(sc_count_list[i])+']</label>'+'<label style="background-color: #DFFF00;"> ['+str(sc_time_list[i])+']</label>')
                    # print(elr_vis_res_list,sc_vis_res_list,elr_nok_indices,sc_nok_indices)
                    html_str=f'''<!DOCTYPE html>
                                    <html>
                                        <body>            
                                            <div   style="border-radius: 10px;
                                                        width:fit-content;
                                                        background-color:rgb(255, 243, 142 ) ;
                                                        height: 100%;
                                                        overflow: auto;
                                                        margin: auto;">
                                            <img src="https://drive.google.com/thumbnail?id=1LMT6hFZXN53ePHpemgdPRhXjM-TLeXNi" style="padding:15px;" width="101" height="42" align="right">
                                            <div style="border-radius: 0px;
                                                    font-family: 'Arial black';
                                                    padding:10px;
                                                    background-color: rgb(25, 111, 61 );
                                                    margin-left: 10px;
                                                    margin-top: 100px;
                                                    height: fit-content;
                                                    width: 200px;
                                                    color: white ;
                                                    font-size: 14px;
                                                    position: relative;
                                                    float: left;
                                                    text-align: left;">
                                                <h3 style ='text-align: center;'>
                                                MSIL Sensor Cover Press Vision Check - Summary</h3>
                                                DATE &nbsp;&nbsp;: {date} <br>
                                                SHIFT &nbsp;: {shift}<br>
                                                FTR &nbsp;&nbsp; &nbsp;: {ftr_c} ({ftr_p}%)
                                            </div>
                                            <div style="margin-left: 200px;
                                                    padding: 10px ;
                                                    padding-left: 50px;
                                                    padding-right: 10px;
                                                    margin-right: 10px;
                                                    margin-top: 70px;
                                                    margin-bottom: 10px;
                                                    background-color: rgb(213, 245, 227 );;
                                                    height: fit-content;
                                                    width: fit-content;
                                                    color: black;
                                                    font-size: 10px;
                                                    text-align: left;">
                                                <h2>ELR RETESTED: 
                                                    <label style="background-color: rgb(25, 111, 61 );
                                                            color: white;
                                                            padding:5px"">
                                                        {len(elr_count_list)-elr_count_list.count(1)-len(elr_nok_list)}
                                                    </label>
                                                </h2>
                                                <table style="border: 2px solid black;
                                                        margin-left: 30px;
                                                        border-collapse:collapse;
                                                        font-size: 15px;
                                                        background-color:rgb(250, 215, 160);">
                                                    <tr>
                                                        <th style="padding:10px;
                                                                text-align: center;
                                                                background-color: white;
                                                                color: black;
                                                                border: 1px solid black;
                                                                white-space: nowrap;
                                                                background-color: orange;
                                                                color: white;">
                                                            ELR Barcode data
                                                        </th>
                                                        <th style="padding:10px;
                                                                text-align: center;
                                                                background-color: white;
                                                                color: black;
                                                                border: none;
                                                                white-space: nowrap;
                                                                background-color: orange;
                                                                color: white;">
                                                            No of times checked
                                                        </th>
                                                    </tr>'''
                    if elr_ok_dic!={}:
                        for count in elr_ok_dic:
                            html_str=html_str+'''<tr align= "center" >
                                                    <td style="padding:10px;
                                                            border: 1px solid black;
                                                            white-space: nowrap;">'''
                            s=''
                            for data in elr_ok_dic[count]:
                                s=s+data+'<br>'
                            html_str=html_str+s+f'''</td>
                                                    <td style="padding:10px;
                                                            border: 1px solid black;
                                                            white-space: nowrap;">
                                                        {count}
                                                    </td>
                                                </tr>'''
                        html_str=html_str+'</table>'
                    else:
                        html_str=html_str+'''<tr align= "center" >
                                                <td style="padding:10px;
                                                        border: 1px solid black;
                                                        white-space: nowrap;">
                                                    No data 
                                                </td>
                                                <td style="padding:10px;
                                                        border: 1px solid black;
                                                        white-space: nowrap;">
                                                    0
                                                </td>
                                            </tr>
                                        </table>'''
                    html_str=html_str+f'''<h2>SENSOR COVER RETESTED: 
                                                <label style="background-color: rgb(25, 111, 61 );
                                                            color: white;
                                                            padding:5px"">
                                                    {len(sc_count_list)-sc_count_list.count(1)-len(sc_nok_list)}
                                                </label>
                                        </h2>
                                        <table style="border: 2px solid black; 
                                                    margin-left: 30px;
                                                    border-collapse:collapse;
                                                    font-size: 15px;
                                                    background-color:rgb(250, 215, 160);">
                                            <tr >
                                                <th style="padding:10px;
                                                        text-align: center;
                                                        background-color: white;
                                                        color: black;border: 1px solid black;
                                                        white-space: nowrap;
                                                        background-color: orange;
                                                        color: white;">
                                                    Sensor Cover Barcode data
                                                </th>
                                                <th style="padding:10px;
                                                        text-align: center;
                                                        background-color: white;
                                                        color: black;
                                                        border: none;
                                                        white-space: nowrap;
                                                        background-color: orange;
                                                        color: white;">
                                                    No of times checked
                                                </th>
                                            </tr>'''
                    if sc_ok_dic!={}:
                        for count in sc_ok_dic:
                            html_str=html_str+'''<tr align= "center" >
                                                    <td style="padding:10px;
                                                            border: 1px solid black;
                                                            white-space: nowrap;">'''
                            s=''
                            for data in sc_ok_dic[count]:
                                s=s+data+'<br>'
                            html_str=html_str+s+f'''</td>
                                                    <td style="padding:10px;
                                                            border: 1px solid black;
                                                            white-space: nowrap;">
                                                        {count}
                                                    </td>
                                                </tr>'''
                        html_str=html_str+'</table>'
                    else:
                        html_str=html_str+'''<tr align= "center" >
                                                <td style="padding:10px;
                                                        border: 1px solid black;
                                                        white-space: nowrap;">
                                                    No data 
                                                </td>
                                                <td style="padding:10px;
                                                        border: 1px solid black;
                                                        white-space: nowrap;">
                                                    0
                                                </td>
                                            </tr>
                                        </table>'''
                    enl=len(elr_nok_list)
                    snl=len(sc_nok_list)
                    html_str=html_str+f'''<h2>FAILED PARTS: <br>
                                            <label style="background-color: rgb(25, 111, 61 );
                                                            color: white;
                                                            padding:5px"">
                                                ELR: {enl}
                                            </label>
                                            &nbsp;
                                            <label style="background-color: rgb(25, 111, 61 );
                                                            color: white;
                                                            padding:5px"">
                                                Sensor Cover: {snl}
                                            </label>
                                        </h2>
                                        <table style="border: 2px solid black;
                                                    margin-left: 30px;
                                                    border-collapse:collapse;
                                                    font-size: 15px;
                                                    background-color:rgb(255, 228, 196);">                
                                            <tr >
                                                <th style="padding:10px;
                                                        text-align: center;
                                                        background-color: white;
                                                        color: black;
                                                        border: 1px solid black;
                                                        white-space: nowrap;
                                                        background-color: red;
                                                        color: white;">
                                                    ELR Barcode data
                                                </th>
                                                <th style="padding:10px;
                                                        text-align: center;
                                                        background-color: white;
                                                        color: black;border: none;
                                                        white-space: nowrap;
                                                        background-color: red;
                                                        color: white;">
                                                    Sensor Cover Barcode data
                                                </th>
                                            </tr>'''
                    if enl!=0:
                        html_str=html_str+'''<tr align= "center" >
                                                <td style="padding:10px;
                                                        border: 1px solid black;
                                                        white-space: nowrap;">'''
                        s=''
                        for i in elr_nok_list:
                            s=s+i+'<br>'
                        html_str=html_str+s+'</td>'
                    else:
                        html_str=html_str+'''<tr align= "center" >
                                                <td style="padding:10px;
                                                        border: 1px solid black;
                                                        white-space: nowrap;">
                                                    No data 
                                                </td>'''
                    if snl!=0:
                        html_str=html_str+'''<td style="padding:10px;
                                                    border: 1px solid black;
                                                    white-space: nowrap;">'''
                        s=''
                        for i in sc_nok_list:
                            s=s+i+'<br>'
                        html_str=html_str+s+'</td></tr>'
                    else:
                        html_str=html_str+'''<td style="padding:10px;
                                                    border: 1px solid black;
                                                    white-space: nowrap;">
                                                No data 
                                            </td>
                                        </tr>'''
                    html_str=html_str+'''</table>  
                                    </body>
                                </html>
                                '''
                    body=html_str
                    message = MIMEMultipart()
                    message['Subject'] = 'MSIL_BB Pin Press Vision report'
                    message['From'] = mail_user_name
                    message['To'] = ', '.join( To_list_of_mails )
                    message['Cc'] = ', '.join( Cc_list_of_mails )
                    body_content = body
                    message.attach(MIMEText(body_content, "html"))
                    msg_body = message.as_string()
                    count=0
                    while True:
                        try:
                            server = SMTP('smtp.gmail.com', 587)
                            server.starttls()
                            server.login(message['From'], app_password)
                            server.send_message(message)
                            server.quit()
                            print("SM: Summary mail sent...")
                        except Exception as e:
                            print("SM: Unable to sent mail, ",e)
                            time.sleep(3)
                            count=count+1
                            if count<3:
                                continue
                            else:
                                break
                        break
                else:
                    print("No data for the shift")
            else:
                print("Excel file does not exist")
            break
        except Exception as e:
            count=count+1
            print(e)
            time.sleep(retry_delay)

def socket_thread():
    ServerSideSocket = socket.socket()
    ThreadCount = 0
    try:
        ServerSideSocket.bind((ipaddress_of_system, port_to_listen))
    except socket.error as e:
        print(str(e))
    ServerSideSocket.listen(5)
    while True:
        print(' MT: Socket is listening..')
        Client, address = ServerSideSocket.accept()
        print('Connected to: ' + address[0] + ':' + str(address[1]))
        client_han = threading.Thread(target=multi_threaded_client,args=(Client,),daemon=True)
        client_han.start()
        ThreadCount += 1
        print('Thread Number: ' + str(ThreadCount))
    # ServerSideSocket.close()

def schedule_thread():
        schedule.every().day.at(shiftA_start).do(make_content_and_send_summary)
        schedule.every().day.at(shiftB_start).do(make_content_and_send_summary)
        schedule.every().day.at(shiftC_start).do(make_content_and_send_summary)
        while True:
            try:            
                schedule.run_pending()
                time.sleep(scheduler_delay)
            except Exception as e :
                print(e)
                time.sleep(3)

socket_th = threading.Thread(target=socket_thread,daemon=True)
socket_th.start()
print("MT: Socket_thread thread started")
move_ex = threading.Thread(target=move_excel,daemon=True)
move_ex.start()
print("MT: Move_excel thread started")
schedule_th = threading.Thread(target=schedule_thread,daemon=True)
schedule_th.start()
print("MT: Schedule_thread thread started")

while True:
    time.sleep(60)